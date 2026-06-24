from datetime import datetime
import json
from pathlib import Path
from typing import List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, Form, Request
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Solicitud
from app.dependencies import get_usuario_actual
from app.permissions import ROLES_SOLICITUDES, puede_editar_datos_solicitante, usuario_tiene_rol
from app.services.catalogo_solicitudes import AREAS_SOLICITUD_ACTIVAS, CATALOGO_SERVICIOS
from app.services.plantilla_solicitud import (
    generar_plantilla_solicitud,
    guardar_plantilla_solicitud,
    nombre_archivo_solicitud_generada,
)
from app.services.solicitudes import (
    AreaSolicitanteInvalida,
    OpcionServicioInvalida,
    SubcategoriaServicioInvalida,
    crear_solicitud,
)


router = APIRouter()

TELEFONO_JEFE_MANTENIMIENTO = "5556228222 ext. 82581"
RUTA_DIRECTORIO_AUTOCOMPLETE = Path(__file__).resolve().parents[2] / "dir_autocomplete.json"
RUTA_SOLICITUDES = Path(__file__).resolve().parents[1] / "solicitudes"


def listar_archivos_xlsx() -> list[dict[str, str]]:
    RUTA_SOLICITUDES.mkdir(parents=True, exist_ok=True)
    archivos = []
    for ruta in sorted(
        RUTA_SOLICITUDES.glob("*.xlsx"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    ):
        estadisticas = ruta.stat()
        archivos.append(
            {
                "nombre": ruta.name,
                "tamano_kb": f"{estadisticas.st_size / 1024:.1f}",
                "modificado": datetime.fromtimestamp(estadisticas.st_mtime).strftime(
                    "%d/%m/%Y %H:%M"
                ),
            }
        )
    return archivos


def ruta_xlsx_segura(nombre_archivo: str) -> Path | None:
    if Path(nombre_archivo).name != nombre_archivo or not nombre_archivo.lower().endswith(
        ".xlsx"
    ):
        return None

    ruta = (RUTA_SOLICITUDES / nombre_archivo).resolve()
    try:
        ruta.relative_to(RUTA_SOLICITUDES.resolve())
    except ValueError:
        return None

    if not ruta.is_file():
        return None
    return ruta


def leer_libro_xlsx(ruta_archivo: Path) -> list[dict[str, object]]:
    libro = load_workbook(ruta_archivo, read_only=True, data_only=True)
    hojas = []
    try:
        for hoja in libro.worksheets:
            filas = []
            for fila in hoja.iter_rows(values_only=True):
                filas.append(["" if celda is None else str(celda) for celda in fila])

            max_columnas = max((len(fila) for fila in filas), default=0)
            hojas.append(
                {
                    "nombre": hoja.title,
                    "encabezados": [
                        get_column_letter(indice)
                        for indice in range(1, max_columnas + 1)
                    ],
                    "filas": filas,
                }
            )
    finally:
        libro.close()

    return hojas


def cargar_personas_autocomplete() -> list[dict[str, str]]:
    with RUTA_DIRECTORIO_AUTOCOMPLETE.open(encoding="utf-8") as archivo:
        directorio = json.load(archivo)

    personas = []
    for persona in directorio.get("personas", []):
        nombre = str(persona.get("nombre", "")).strip()
        if not nombre:
            continue

        etiqueta = str(persona.get("etiqueta", "")).strip()
        personas.append({"nombre": nombre, "etiqueta": etiqueta})

    return personas


def contexto_formulario(usuario: dict, **valores) -> dict:
    puede_editar = puede_editar_datos_solicitante(usuario)
    return {
        "areas": AREAS_SOLICITUD_ACTIVAS,
        "catalogo_servicios": CATALOGO_SERVICIOS,
        "nombre_usuario": valores.get("nombre_usuario", usuario.get("nombre", "")),
        "responsable_area_solicitante": valores.get("responsable_area_solicitante", ""),
        "telefono": valores.get("telefono", TELEFONO_JEFE_MANTENIMIENTO),
        "area_solicitante": valores.get("area_solicitante", ""),
        "subcategoria_servicio": valores.get("subcategoria_servicio", ""),
        "descripcion_servicio": valores.get("descripcion_servicio", ""),
        "personas_autocomplete": cargar_personas_autocomplete(),
        "puede_editar_datos_solicitante": puede_editar,
    }


def contexto_confirmacion(solicitud: Solicitud, **valores) -> dict:
    return {
        "folio": solicitud.folio,
        "fecha": solicitud.fecha.strftime("%d/%m/%Y"),
        "nombre_usuario": solicitud.nombre_usuario,
        "responsable_area_solicitante": solicitud.responsable_area_solicitante,
        "telefono": solicitud.telefono,
        "area_solicitante": solicitud.area_solicitante,
        "url_guardar_plantilla": f"/solicitud/{solicitud.folio}/plantilla/guardar",
        **valores,
    }


@router.get("/solicitudes", response_class=HTMLResponse)
async def listar_solicitudes_guardadas(
    request: Request,
    usuario: dict = Depends(get_usuario_actual),
):
    if not usuario_tiene_rol(usuario, ROLES_SOLICITUDES):
        return RedirectResponse(url="/login", status_code=302)

    templates = request.app.state.templates
    return templates.TemplateResponse(
        request=request,
        name="solicitudes_xlsx.html",
        context={"archivos": listar_archivos_xlsx()},
    )


@router.get("/solicitudes/{nombre_archivo}", response_class=HTMLResponse)
async def ver_solicitud_guardada(
    nombre_archivo: str,
    request: Request,
    usuario: dict = Depends(get_usuario_actual),
):
    if not usuario_tiene_rol(usuario, ROLES_SOLICITUDES):
        return RedirectResponse(url="/login", status_code=302)

    ruta_archivo = ruta_xlsx_segura(nombre_archivo)
    templates = request.app.state.templates
    if ruta_archivo is None:
        return templates.TemplateResponse(
            request=request,
            name="solicitudes_xlsx.html",
            context={
                "archivos": listar_archivos_xlsx(),
                "error": "No se encontró el archivo XLSX solicitado.",
            },
            status_code=404,
        )

    return templates.TemplateResponse(
        request=request,
        name="visor_xlsx.html",
        context={
            "archivo": ruta_archivo.name,
            "hojas": leer_libro_xlsx(ruta_archivo),
            "url_descarga": f"/solicitudes/{quote(ruta_archivo.name)}/descargar",
        },
    )


@router.get("/solicitudes/{nombre_archivo}/descargar")
async def descargar_solicitud_guardada(
    nombre_archivo: str,
    usuario: dict = Depends(get_usuario_actual),
):
    if not usuario_tiene_rol(usuario, ROLES_SOLICITUDES):
        return RedirectResponse(url="/login", status_code=302)

    ruta_archivo = ruta_xlsx_segura(nombre_archivo)
    if ruta_archivo is None:
        return RedirectResponse(url="/solicitudes", status_code=302)

    return FileResponse(
        path=ruta_archivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=ruta_archivo.name,
    )


@router.get("/solicitud", response_class=HTMLResponse)
async def mostrar_formulario(
    request: Request,
    usuario: dict = Depends(get_usuario_actual),
):
    if not usuario_tiene_rol(usuario, ROLES_SOLICITUDES):
        return RedirectResponse(url="/login", status_code=302)

    templates = request.app.state.templates
    return templates.TemplateResponse(
        request=request,
        name="form.html",
        context=contexto_formulario(usuario),
    )


@router.get("/solicitud/{folio}/plantilla")
async def descargar_plantilla_solicitud(
    folio: str,
    db: Session = Depends(get_db),
    usuario: dict = Depends(get_usuario_actual),
):
    if not usuario_tiene_rol(usuario, ROLES_SOLICITUDES):
        return RedirectResponse(url="/login", status_code=302)

    solicitud = db.query(Solicitud).filter(Solicitud.folio == folio).first()
    if solicitud is None:
        return RedirectResponse(url="/solicitud", status_code=302)

    contenido = generar_plantilla_solicitud(solicitud)
    nombre_descarga = nombre_archivo_solicitud_generada(solicitud)
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_descarga}"'},
    )


@router.post("/solicitud/{folio}/plantilla/guardar", response_class=HTMLResponse)
async def guardar_plantilla_solicitud_en_servidor(
    folio: str,
    request: Request,
    db: Session = Depends(get_db),
    usuario: dict = Depends(get_usuario_actual),
):
    if not usuario_tiene_rol(usuario, ROLES_SOLICITUDES):
        return RedirectResponse(url="/login", status_code=302)

    solicitud = db.query(Solicitud).filter(Solicitud.folio == folio).first()
    if solicitud is None:
        return RedirectResponse(url="/solicitud", status_code=302)

    ruta_archivo = guardar_plantilla_solicitud(solicitud)
    templates = request.app.state.templates
    return templates.TemplateResponse(
        request=request,
        name="confirmacion.html",
        context=contexto_confirmacion(
            solicitud,
            archivo_guardado=ruta_archivo.relative_to(
                Path(__file__).resolve().parents[2]
            ),
        ),
    )


@router.post("/solicitud", response_class=HTMLResponse)
async def recibir_formulario(
    request: Request,
    db: Session = Depends(get_db),
    usuario: dict = Depends(get_usuario_actual),
    area_solicitante: str = Form(...),
    nombre_usuario: str = Form(...),
    responsable_area_solicitante: str = Form(...),
    telefono: str = Form(TELEFONO_JEFE_MANTENIMIENTO),
    descripcion_servicio: str = Form(...),
    subcategoria_servicio: Optional[str] = Form(None),
    infraestructura: Optional[List[str]] = Form(None),
    equipo_parque_vehicular: Optional[List[str]] = Form(None),
    seguridad: Optional[List[str]] = Form(None),
    transporte: Optional[List[str]] = Form(None),
    diversos_limpieza: Optional[List[str]] = Form(None),
    prestamo_de: Optional[List[str]] = Form(None),
    correspondencia_paqueteria: Optional[List[str]] = Form(None),
    reproduccion_engargolado: Optional[List[str]] = Form(None),
):
    if not usuario_tiene_rol(usuario, ROLES_SOLICITUDES):
        return RedirectResponse(url="/login", status_code=302)

    templates = request.app.state.templates

    nombre_usuario = nombre_usuario.strip()
    responsable_area_solicitante = responsable_area_solicitante.strip()
    telefono = telefono.strip() or TELEFONO_JEFE_MANTENIMIENTO

    if not puede_editar_datos_solicitante(usuario):
        nombre_usuario = str(usuario.get("nombre", "")).strip()
        responsable_area_solicitante = responsable_area_solicitante or nombre_usuario

    try:
        solicitud = crear_solicitud(
            db,
            nombre_usuario=nombre_usuario,
            telefono=telefono,
            responsable_area_solicitante=responsable_area_solicitante,
            area_solicitante=area_solicitante,
            descripcion_servicio=descripcion_servicio,
            subcategoria_servicio=subcategoria_servicio,
            infraestructura=infraestructura,
            equipo_parque_vehicular=equipo_parque_vehicular,
            seguridad=seguridad,
            transporte=transporte,
            diversos_limpieza=diversos_limpieza,
            prestamo_de=prestamo_de,
            correspondencia_paqueteria=correspondencia_paqueteria,
            reproduccion_engargolado=reproduccion_engargolado,
        )
    except (AreaSolicitanteInvalida, SubcategoriaServicioInvalida, OpcionServicioInvalida) as exc:
        return templates.TemplateResponse(
            request=request,
            name="form.html",
            context={
                **contexto_formulario(
                    usuario,
                    nombre_usuario=nombre_usuario,
                    responsable_area_solicitante=responsable_area_solicitante,
                    telefono=telefono,
                    area_solicitante=area_solicitante,
                    subcategoria_servicio=subcategoria_servicio,
                    descripcion_servicio=descripcion_servicio,
                ),
                "error": str(exc),
            },
            status_code=400,
        )

    return templates.TemplateResponse(
        request=request,
        name="confirmacion.html",
        context=contexto_confirmacion(solicitud),
    )
