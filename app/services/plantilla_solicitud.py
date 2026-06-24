"""Generación de copias llenas de la plantilla de solicitud única.

Enfoque: edición directa del XML de la hoja, sin regenerar el archivo.

A diferencia de usar openpyxl para escribir (que regenera la hoja y, al hacerlo,
altera o descarta los dibujos: imágenes y autoformas como las líneas doradas),
aquí se edita el `sheet1.xml` original reemplazando solo el contenido de las
celdas necesarias, y se vuelve a empaquetar el .xlsx copiando TODO lo demás byte
a byte. Así el dibujo de la plantilla (imágenes y líneas) queda intacto, porque
nunca se reescribe.

Las celdas se escriben como `inlineStr`, que guardan el texto dentro de la
propia celda y no dependen de la tabla de cadenas compartidas
(`sharedStrings.xml`). El estilo de cada celda (fuente, bordes, relleno) se
conserva leyendo su atributo `s` del XML original.
"""

import re
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, Tuple, Union
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.cell import coordinate_to_tuple

from app.models import Solicitud

RAIZ_PROYECTO = Path(__file__).resolve().parents[2]
RUTA_PLANTILLA_SOLICITUD = RAIZ_PROYECTO / "plantilla_solicitud_unica_servicios_pabellon.xlsx"

MARCA_OPCION = "X"
NOMBRE_ARCHIVO_SOLICITUD = "plantilla_solicitud_unica_de_servicios.xlsx"
RUTA_SOLICITUDES_GENERADAS = RAIZ_PROYECTO / "app" / "solicitudes"


# Ruta interna del XML de la hoja dentro del paquete .xlsx.
RUTA_HOJA_XML = "xl/worksheets/sheet1.xml"

# Ruta interna de los estilos. Se ajusta para centrar verticalmente la
# descripción del servicio (celda B37).
RUTA_ESTILOS_XML = "xl/styles.xml"

# El estilo de la celda de descripción (B37) ya viene centrado horizontalmente
# pero alineado arriba. Para centrarlo también en vertical se cambia, solo en
# ese estilo, vertical="top" por vertical="center". Se busca la firma completa
# del estilo para no afectar a ningún otro.
FIRMA_ESTILO_DESCRIPCION = (
    '<xf numFmtId="0" fontId="11" fillId="0" borderId="22" xfId="0" '
    'applyFont="1" applyBorder="1" applyAlignment="1">'
    '<alignment horizontal="center" vertical="top"/>'
)
FIRMA_ESTILO_DESCRIPCION_CENTRADO = FIRMA_ESTILO_DESCRIPCION.replace(
    'vertical="top"', 'vertical="center"'
)

# Las marcas se colocan en la celda indicada para cada opción seleccionada.
CELDAS_OPCIONES_SERVICIO: Dict[str, Dict[str, str]] = {
    "infraestructura": {
        "albanileria": "F17",
        "carpinteria": "F19",
        "electricidad": "F21",
        "herreria": "F23",
        "pintura": "K17",
        "plomeria": "K19",
        "otro": "K21",
    },
    "equipo_parque_vehicular": {
        "mecanica": "R17",
        "refrigeracion": "R19",
        "aire_acondicionado": "R21",
        "equipo_computo": "R23",
        "reparacion_equipo": "X17",
        "planta_luz": "X19",
        "otro": "X21",
    },
    "seguridad": {
        "vigilancia_eventos": "AE17",
        "control_accesos": "AE20",
        "otro": "AE23",
    },
    "transporte": {
        "local": "F28",
        "foraneo": "F30",
        "pasajeros": "F32",
        "carga": "F34",
    },
    "prestamo_de": {
        "salas_aulas": "N30",
        "auditorio": "N32",
        "equipo_audiovisual": "N34",
    },
    "diversos_limpieza": {
        "cafeteria": "R28",
        "cerrajeria": "R30",
        "limpieza": "R32",
        "otro": "R34",
    },
    "correspondencia_paqueteria": {
        "propio": "X30",
        "correo_ordinario": "X32",
        "mensajeria_especializada": "X34",
    },
    "reproduccion_engargolado": {
        "reproduccion": "AE30",
        "engargolado": "AE32",
        "otro": "AE34",
    },
}

CAMPOS_OPCIONES: Tuple[str, ...] = tuple(CELDAS_OPCIONES_SERVICIO.keys())
ValorCelda = Union[int, str]


def _opciones_seleccionadas(solicitud: Solicitud, campo: str) -> Iterable[str]:
    return getattr(solicitud, campo) or []


def _resolver_celda_ancla(hoja, referencia: str) -> str:
    """Devuelve la celda superior-izquierda del rango combinado que contiene
    `referencia`. Si la celda no está combinada, devuelve la misma referencia.

    En un rango combinado solo la celda ancla guarda el valor; escribir en otra
    celda del rango lo pierde. Resolver la ancla aquí mantiene el código
    correcto aunque a futuro se combinen más celdas.
    """
    fila, columna = coordinate_to_tuple(referencia)
    for rango in hoja.merged_cells.ranges:
        min_col, min_fila, max_col, max_fila = range_boundaries(str(rango))
        if min_fila <= fila <= max_fila and min_col <= columna <= max_col:
            return f"{get_column_letter(min_col)}{min_fila}"
    return referencia


def _mapa_valores(solicitud: Solicitud) -> Dict[str, ValorCelda]:
    """Construye el diccionario celda -> valor con las referencias ya resueltas
    a su celda ancla (por si alguna cae en un rango combinado)."""
    # Se abre la plantilla solo para conocer los rangos combinados y resolver
    # anclas; no se usa openpyxl para escribir.
    libro = load_workbook(RUTA_PLANTILLA_SOLICITUD)
    hoja = libro[libro.sheetnames[0]]

    crudos: Dict[str, ValorCelda] = {
        "H7": solicitud.area_solicitante,
        "L9": solicitud.nombre_usuario,
        "AD9": solicitud.fecha.day,
        "AE9": solicitud.fecha.month,
        # El año se muestra con dos dígitos (p. ej. 26 en lugar de 2026).
        "AF9": solicitud.fecha.year % 100,
        "I11": solicitud.responsable_area_solicitante or solicitud.nombre_usuario,
        "AC11": solicitud.telefono,
        # La descripción cae en el bloque combinado B37:AF41; la ancla es B37.
        "B37": solicitud.descripcion_servicio,
        "U56": solicitud.nombre_usuario,
    }
    # El folio (AC7) se deja vacío a propósito: lo asigna un departamento externo
    # a la aplicación, así que la plantilla no debe rellenarlo.

    for campo in CAMPOS_OPCIONES:
        celdas_por_opcion = CELDAS_OPCIONES_SERVICIO[campo]
        for opcion in _opciones_seleccionadas(solicitud, campo):
            celda = celdas_por_opcion.get(opcion)
            if celda:
                crudos[celda] = MARCA_OPCION

    # Resolver anclas y descartar valores None.
    valores: Dict[str, ValorCelda] = {}
    for referencia, valor in crudos.items():
        if valor is None:
            continue
        ancla = _resolver_celda_ancla(hoja, referencia)
        valores[ancla] = valor

    return valores


def _estilo_celda(hoja_xml: str, referencia: str) -> Union[str, None]:
    """Devuelve el índice de estilo (atributo s) de la celda en el XML, o None."""
    coincidencia = re.search(rf'<c r="{referencia}"(?:\s+s="(\d+)")?', hoja_xml)
    if coincidencia and coincidencia.group(1):
        return coincidencia.group(1)
    return None


def _xml_celda(referencia: str, estilo: Union[str, None], valor: ValorCelda) -> str:
    """Construye el XML de una celda, conservando su estilo. Los números se
    escriben como valor numérico; el texto, como cadena en línea (inlineStr)."""
    atributo_estilo = f' s="{estilo}"' if estilo else ""
    if isinstance(valor, int):
        return f'<c r="{referencia}"{atributo_estilo}><v>{valor}</v></c>'
    texto = escape(str(valor))
    return (
        f'<c r="{referencia}"{atributo_estilo} t="inlineStr">'
        f'<is><t xml:space="preserve">{texto}</t></is></c>'
    )


def _reemplazar_celda(hoja_xml: str, referencia: str, valor: ValorCelda) -> str:
    """Reemplaza la celda `referencia` en el XML por una con el nuevo valor,
    conservando su estilo. Todas las celdas destino existen en la plantilla, así
    que siempre se reemplaza (no se inserta)."""
    estilo = _estilo_celda(hoja_xml, referencia)
    nueva_celda = _xml_celda(referencia, estilo, valor)
    patron = rf'<c r="{referencia}"[^>]*?/>|<c r="{referencia}"[^>]*?>.*?</c>'
    if re.search(patron, hoja_xml):
        return re.sub(patron, nueva_celda, hoja_xml, count=1)
    return hoja_xml


def _centrar_descripcion(estilos_xml: str) -> str:
    """Centra verticalmente el texto de la celda de descripción (B37) ajustando
    su estilo en styles.xml. El centrado horizontal ya viene de la plantilla."""
    return estilos_xml.replace(
        FIRMA_ESTILO_DESCRIPCION, FIRMA_ESTILO_DESCRIPCION_CENTRADO
    )


def generar_plantilla_solicitud(solicitud: Solicitud) -> bytes:
    """Devuelve una copia XLSX de la plantilla llena con los datos de la
    solicitud, conservando intactos los dibujos (imágenes y líneas)."""
    valores = _mapa_valores(solicitud)

    with zipfile.ZipFile(RUTA_PLANTILLA_SOLICITUD, "r") as plantilla:
        hoja_xml = plantilla.read(RUTA_HOJA_XML).decode("utf-8")
        for referencia, valor in valores.items():
            hoja_xml = _reemplazar_celda(hoja_xml, referencia, valor)

        estilos_xml = plantilla.read(RUTA_ESTILOS_XML).decode("utf-8")
        estilos_xml = _centrar_descripcion(estilos_xml)

        salida = BytesIO()
        with zipfile.ZipFile(salida, "w", zipfile.ZIP_DEFLATED) as resultado:
            for nombre in plantilla.namelist():
                if nombre == RUTA_HOJA_XML:
                    resultado.writestr(nombre, hoja_xml.encode("utf-8"))
                elif nombre == RUTA_ESTILOS_XML:
                    resultado.writestr(nombre, estilos_xml.encode("utf-8"))
                else:
                    # Todo lo demás (dibujos, imágenes, relaciones) se copia sin
                    # modificar, preservando el diseño original.
                    resultado.writestr(nombre, plantilla.read(nombre))

    return salida.getvalue()


def nombre_archivo_solicitud_generada(solicitud: Solicitud) -> str:
    """Devuelve el nombre del archivo XLSX generado para una solicitud."""
    return f"{solicitud.folio}_{NOMBRE_ARCHIVO_SOLICITUD}"


def guardar_plantilla_solicitud(solicitud: Solicitud) -> Path:
    """Genera y guarda la plantilla XLSX en la carpeta interna de solicitudes."""
    RUTA_SOLICITUDES_GENERADAS.mkdir(parents=True, exist_ok=True)
    ruta_archivo = (
        RUTA_SOLICITUDES_GENERADAS / nombre_archivo_solicitud_generada(solicitud)
    )
    ruta_archivo.write_bytes(generar_plantilla_solicitud(solicitud))
    return ruta_archivo
